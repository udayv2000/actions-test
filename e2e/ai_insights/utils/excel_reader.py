# excel_reader.py

from openpyxl import load_workbook
from pathlib import Path

# Define the file path once at the top
EXCEL_FILE_PATH = r"D:\automation-lattice\e2e\ai_insights\testData\testdata.xlsx"  # ← Change this to your actual file path

def read_test_data(sheet_name, parent_tcid, data_tcid):
    """
    Universal test data reader.
    For ANY parent_tcid and data_tcid:
        - parent_tcid = header row (defines field names)
        - data_tcid  = data row (contains values)

    Returns: dict {field_name: value}
    Example: {'url': 'abc.com', 'appname': 'flipkart', ...}
    """
    workbook = None
    try:
        # Resolve file path
        file_path = Path(EXCEL_FILE_PATH).resolve()
        if not file_path.exists():
            raise FileNotFoundError(f"Test data file not found: {file_path}")

        workbook = load_workbook(filename=file_path, data_only=True)

        # Select sheet
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        worksheet = workbook[sheet_name]

        # 🔹 Step 1: Read headers from parent_tcid row
        parent_row = None
        for row in worksheet.iter_rows():
            if row[0].value == parent_tcid:
                parent_row = row
                break
        if not parent_row:
            raise ValueError(f"Header row not found for parent TCID: '{parent_tcid}'")

        headers = []
        for cell in parent_row[1:]:  # Skip TCID column
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            # Skip empty columns (no dummy names)

        # 🔹 Step 2: Read data from data_tcid row
        data_row = None
        for row in worksheet.iter_rows():
            if row[0].value == data_tcid:
                data_row = row
                break
        if not data_row:
            raise ValueError(f"Data row not found for data TCID: '{data_tcid}'")

        # 🔹 Step 3: Map headers to values
        data = {}
        for i, header in enumerate(headers):
            if i + 1 < len(data_row):  # Check if value exists
                value_cell = data_row[i + 1]
                value = value_cell.value if value_cell is not None else ""
                data[header] = str(value).strip()

        return data

    except Exception as e:
        print(f"❌ Error reading test data: {e}")
        raise

    finally:
        if workbook:
            workbook.close()